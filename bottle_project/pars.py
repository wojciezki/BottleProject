import datetime
import csv
import json
import StringIO
import wtforms
import xlsxwriter
import openpyxl
from bottle import request, route, run, view, static_file, redirect, response, FileUpload
from peewee import SqliteDatabase, Model, CharField, DateTimeField


db = SqliteDatabase("data.db")


class Client(Model):
    email = CharField(unique=True)
    first_name = CharField()
    last_name = CharField()
    date = DateTimeField(default=datetime.datetime.now)

    class Meta:
        database = db


class User(Model):
    login = CharField(unique=True)
    password = CharField()
    creation_date = DateTimeField(default=datetime.datetime.now)

    class Meta:
        database = db


class ClientForm(wtforms.Form):
    email = wtforms.StringField('Email', [wtforms.validators.Email()])
    first_name = wtforms.StringField('Name', [wtforms.validators.Length(min=3, max=30)])
    last_name = wtforms.StringField('Surname', [wtforms.validators.Length(min=3, max=50)])


class UserForm(wtforms.Form):
    login = wtforms.StringField('Login', [wtforms.validators.InputRequired()])
    password = wtforms.StringField('Password', [wtforms.validators.InputRequired()])


class UploadForm(wtforms.Form):
    file = wtforms.FileField()

    def validate_image(form, field):
        if field.data:
            field.data = re.sub(r'[^a-z0-9_.-]', '_', field.data)


def get_data_from_file():
    with open("/home/wojtas/PycharmProjects/untitled/wojtas.csv", "rU") as f:
        reader = csv.DictReader(f)
        naglowki = []
        for row in reader:
            naglowki.append(row)
        return naglowki


def json_stuff(dane):
    return json.dumps(dane)


@route("/")
@view("index")
def index():
    return (dict(users=[dict(
        id=i.id,
        first_name=i.first_name,
        last_name=i.last_name,
        email=i.email,
        date=i.date.strftime("%d.%m.%Y %H:%M")
    ) for i in Client.select().order_by(Client.date)]))


@route("/upload/file", method=['GET', 'POST'])
@view("upload")
def upload_file():
    """
    uploading data from xlsx file
    :return:
    """
    form = UploadForm()
    if request.method == 'POST':
        form = UploadForm(request.POST)
        if form.file.data:
            wb = openpyxl.load_workbook(request.files[form.file.name].file)
            ws = wb.active
            row_list = []
            for row in ws.iter_rows():
                row_list.append([cell.value for cell in row])
            keys = row_list[0]
            data = [dict(zip(keys, values)) for values in row_list[1:]]
            for i in data:
                new = Client.create(
                    email=i['email'],
                    first_name=i['first_name'],
                    last_name=i['last_name']
                )
                new.save()
        redirect('/')
    return {'form': form}


@route("/<id:int>/")
@view("row")
def row_stuff(id):
    row = Client.select().where(Client.id == id).get()
    return {
        'user': {
            'id': row.id,
            'first_name': row.first_name,
            'last_name': row.last_name,
            'email': row.email,
            'date': row.date.strftime("%d.%m.%Y %H:%M"),
        }
    }


@route("/<id:int>/edit/", method=["GET", "POST"])
@view("edit")
def edit_data(id):
    import pdb
    pdb.set_trace()
    client = Client.select().where(Client.id == id).get()
    form = ClientForm(obj=client)
    if request.method == 'POST':
        form = ClientForm(request.POST, client)
        if form.validate():
            form.populate_obj(client)
            client.save()
            redirect("/{id}/".format(id=id))
    return {'form': form, "id": id}


@route("/create_data/", method=["GET", "POST"])
@view("create_data")
def create_data():
    form = ClientForm()
    if request.method == 'POST':
        form = ClientForm(request.POST)
        if form.validate():
            new = Client.create(
                email=form.email.data,
                first_name=form.first_name.data,
                last_name=form.last_name.data
            )
            new.save()
            redirect("/")
    return {'form': form}


@route("/<id:int>/delete/")
def delete_data(id):
    client = Client.get(Client.id == id)
    client.delete_instance()
    redirect("/")


@route("/users_list/")
@view("users")
def user_list():
    return dict(users_list=[dict(
        id=i.id,
        login=i.login,
        password=i.password,
        creation_date=i.creation_date.strftime("%d.%m.%Y %H:%M")
    ) for i in User.select()])


@route("/create_user/", method=["GET", "POST"])
@view("create_user")
def create_user():
    form = UserForm()
    if request.method == 'POST':
        form = UserForm(request.POST)
        if form.validate():
            new = User.create(
                login=form.login.data,
                password=form.password.data
            )
            new.save()
            redirect('/users_list/')
    return {"form": form}


@route("/<id:int>/edit_user/", method=['POST', 'GET'])
@view("edit_user")
def edit_data(id):
    user = User.select().where(User.id == id).get()
    form = UserForm(obj=user)
    if request.method == 'POST':
        form = UserForm(request.POST, user)
        if form.validate():
            form.populate_obj(user)
            user.save()
            redirect("/users_list/")
    return {'form': form, "id": id}


@route("/<id:int>/user_delete/")
def user_delete(id):
    user = User.get(User.id == id)
    user.delete_instance()
    redirect("/users_list/")


@route("/api/client/", method=["GET"])
def get_client_list():
    return json.dumps([dict(
        id=i.id,
        first_name=i.first_name,
        last_name=i.last_name,
        email=i.email,
        date=i.date.strftime("%d.%m.%Y %H:%M")
    ) for i in Client.select().order_by(Client.last_name)])


@route("/api/client/<id:int>/", method=["GET"])
def get_single_client(id):
    row = Client.select().where(Client.id == id).get()
    return json.dumps({
        'id': row.id,
        'first_name': row.first_name,
        'last_name': row.last_name,
        'email': row.email,
        'date': row.date.strftime("%d.%m.%Y %H:%M"),
    })


@route("/api/client/", method=["POST"])
def adding_client():
    data = request.json
    form = ClientForm()
    if request.method == 'POST':
        form = ClientForm(**data)
        if form.validate():
            new = Client.create(
                email=form.email.data,
                first_name=form.first_name.data,
                last_name=form.last_name.data
            )
            new.save()
        else:
            return form.errors
    return data


@route("/<id:int>/api/client_delete/", method=["GET"])
def client_deleting(id):
    client = Client.get(Client.id == id)
    client.delete_instance()


@route("/export/xlsx/", method=["GET"])
def export_xlsx():
    """
    exporting xls file from web

    :return:
    """
    f = StringIO.StringIO()
    book = xlsxwriter.Workbook(f, {'in_memory': True})
    sheet1 = book.add_worksheet("Python Sheet 1")
    for offset1, i in enumerate(Client):
        dane = dict(
            date=i.date.strftime("%d.%m.%Y %H:%M"),
            first_name=i.first_name,
            last_name=i.last_name,
            email=i.email,
            id=i.id
        )
        for offset2, x in enumerate(dane.values()):
            sheet1.write(offset1 + 1, offset2, x)
    for offset, key in enumerate(dane.keys()):
        sheet1.write(0, offset, key)
    book.close()
    f.seek(0)
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = 'attachment; filename=All.xlsx'
    return f.read()


@route("/export/xlsx/<id:int>/", method=["GET"])
def export_xlsx_client(id):
    """
    Exporting single client
    :param id:
    :return:
    """
    row = 0
    f = StringIO.StringIO()
    client_book = xlsxwriter.Workbook(f, {'in_memory': True})
    client_sheet = client_book.add_worksheet("Python Sheet 1")
    client = Client.select().where(Client.id == id).get()
    dane = {
        'date': client.date.strftime("%d.%m.%Y %H:%M"),
        'first_name': client.first_name,
        'last_name': client.last_name,
        'email': client.email,
        'id': client.id
    }
    for offset, key in enumerate(dane.keys()):
        client_sheet.write(0, offset, key)
    for offset, x in enumerate(dane.values()):
        client_sheet.write(row + 1, offset, x)
    client_book.close()
    f.seek(0)
    response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    response.headers['Content-Disposition'] = 'attachment; filename=Single_Client.xlsx'
    return f.read()


@route('/static/<filepath:path>')
def server_static(filepath):
    return static_file(filepath, root='/home/wojjak/PycharmProjects/untitled/static')


if __name__ == "__main__":
    run(host="localhost", port=8080, debug=True)

